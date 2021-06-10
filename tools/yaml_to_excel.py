import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image

"""
Recommended LibreOffice
"""

align_center = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)
align_left = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0, justifyLastLine=True)

from openpyxl.styles import NamedStyle, Font, Border, Side

def create_style(params):

    style = NamedStyle(name=params.get('name'))
    if params.get('font'):
        style.font = Font(name=params.get('font'))
    # bd = Side(style='thick', color="FFFFFF")
    # style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    alignment = params.get('alignment')
    if alignment == 'center':
        style.alignment = align_center
    else:
        style.alignment = align_left

    if params.get('background'):
        style.fill = PatternFill("solid", fgColor=params.get('background'))
        bd = Side(style='thick', color="FFFFFF")
        style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    return style

# text = NamedStyle(name="text")
# text.font = Font(name='Calibri')
# bd = Side(style='thick', color="FFFFFF")
# text.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# text.alignment = align_left
# styles.append(text)

# gray = NamedStyle(name="gray")
# gray.font = Font(name='Calibri')
# gray.fill = PatternFill("solid", fgColor="00C0C0C0")
# gray.alignment = align_center
# styles.append(gray)

# blue = NamedStyle(name="blue")
# blue.fill = PatternFill("solid", fgColor="0033CCCC")
# blue.alignment = align_center
# styles.append(blue)


# green = NamedStyle(name="green")
# green.fill = PatternFill("solid", fgColor="0099CC00")
# green.alignment = align_center
# styles.append(green)


# orange = NamedStyle(name="orange")
# orange.fill = PatternFill("solid", fgColor="00FF9900")
# orange.alignment = align_center
# styles.append(orange)

# TODO parametrize & create a function
# TODO read sheet, col ids and cell styles (centered, left, colored, character if true, etc) from yaml

MASVS_TITLES = {
    'V1': 'Architecture, Design and Threat Modeling Requirements',
    'V2': 'Data Storage and Privacy Requirements',
    'V3': 'Cryptography Requirements',
    'V4': 'Authentication and Session Management Requirements',
    'V5': 'Network Communication Requirements',
    'V6': 'Platform Interaction Requirements',
    'V7': 'Code Quality and Build Setting Requirements',
    'V8': 'Resilience Requirements',
}

def get_hyperlink(url):
    # title = url.split('#')[1].replace('-',' ').capitalize() 
    if '/0x05' in url:
        title = 'Android'
    elif '/0x06' in url:
        title = 'iOS'
    return f'=HYPERLINK("{url}", "{title}")'

# CHECKMARK = "✓"

def write_table(masvs_file, output_file):

    masvs_dict = yaml.safe_load(open(masvs_file))

    # wb = load_workbook(filename=input_file)

    wb = Workbook()
    # output_file = 'checklist_new.xlsx'
    table = wb.active
    table.title = 'Security Requirements'

    table_config = {
        'sheet': 'Security Requirements - Android',
        'styles': [
            {'name': 'text', 'font': 'Calibri', 'alignment': 'left', 'background': ''},
            {'name': 'center', 'font': 'Calibri', 'alignment': 'center', 'background': ''},
            {'name': 'gray', 'font': 'Calibri', 'alignment': 'center', 'background': '00C0C0C0'},
            {'name': 'blue', 'font': 'Calibri', 'alignment': 'center', 'background': '0033CCCC'},
            {'name': 'green', 'font': 'Calibri', 'alignment': 'center', 'background': '0099CC00'},
            {'name': 'orange', 'font': 'Calibri', 'alignment': 'center', 'background': '00FF9900'},
        ],
        'start_row': 5,
        'start_col': 2,
        'columns': [
            {'name': 'ID', 'width': 10,},
            {'name': 'MSTG-ID', 'width': 25,},
            {'name': 'Detailed Verification Requirement',  'width': 80,},
            {'name': 'L1', 'style': 'blue', 'width': 5,},
            {'name': 'L2', 'style': 'green', 'width': 5,},
            {'name': 'R', 'style': 'orange', 'width': 5,},
            {'name': 'References', 'width': 70,},
        ]
            
    }

    [wb.add_named_style(create_style(style)) for style in table_config.get('styles')]
    
    underline = NamedStyle(name="underline")
    underline.font = Font(name='Calibri', size=15, bold=True, color='00C0C0C0')
    bd = Side(style='medium', color="00C0C0C0")
    underline.border = Border(bottom=bd)
    # underline.alignment = align_center
    wb.add_named_style(underline)

    big_title = NamedStyle(name="big_title")
    big_title.font = Font(name='Calibri', size=25)
    big_title.alignment = align_left
    wb.add_named_style(big_title)

    gray_header = NamedStyle(name="gray_header")
    gray_header.font = Font(name='Calibri', bold=True, color="00C0C0C0")
    # gray_header.fill = PatternFill("solid", fgColor="00C0C0C0")
    gray_header.alignment = align_center
    wb.add_named_style(gray_header)

    # table = wb['Security Requirements']

    table.row_dimensions[2].height = 65
    img = Image('../Document/images/OWASP_logo.png')
    img.height = img.height * 0.08
    img.width = img.width * 0.08
    table.add_image(img, 'B2')

    table['D2'].value = "Mobile Application Security Verification Standard"
    table['D2'].style = big_title

    table['D3'].value = f'=HYPERLINK("https://github.com/OWASP/owasp-masvs/releases/tag/v1.3", "v1.3 (git: 496d39821f17e92fca49eadb35eddb7468012b3c)")'
    table['D3'].font = Font(name='Calibri', color="00C0C0C0")

    table.column_dimensions['B'].width = 5
    table.column_dimensions['C'].width = 23
    table.column_dimensions['D'].width = 80
    table.column_dimensions['E'].width = 5
    table.column_dimensions['F'].width = 5
    table.column_dimensions['G'].width = 5
    table.column_dimensions['H'].width = 10
    table.column_dimensions['I'].width = 10

    row=4
    col_id=2
    col_mstg_id=3
    col_text=4
    col_l1=5
    col_l2=6
    col_r=7
    col_link_android=8
    col_link_ios=9

    for mstg_id, req in masvs_dict.items():
        req_id = req['id'].split('.') 
        category = req_id[0]
        subindex = req_id[1]

        if subindex == '1':
            row = row+1

            category_id = f"V{category}"
            category_title = MASVS_TITLES[category_id]

            # category = f"{category_id} - {category_title}"
            
            category_cell = table.cell(row=row,column=col_id)
            category_cell.value = category_title
            category_cell.style = 'underline'
            category_cell.alignment = align_left

            # title_cell_1 = table.cell(row=row,column=col_id+1)
            # title_cell_1.style = 'underline'
            # title_cell_1.alignment = align_left
            # title_cell_1.value = category_title

            # title_cell_2 = table.cell(row=row,column=col_id+2)
            # title_cell_2.style = 'underline'
            # title_cell_2.alignment = align_left
            

            table.merge_cells(start_row=row, end_row=row, start_column=col_id, end_column=col_link_ios)

            table.row_dimensions[row].height = 25 # points
            row = row+2

            table.cell(row=row,column=col_id).value = 'ID'
            table.cell(row=row,column=col_id).style = 'gray_header'

            table.cell(row=row,column=col_mstg_id).value = 'MSTG-ID'
            table.cell(row=row,column=col_mstg_id).style = 'gray_header'
            
            table.cell(row=row,column=col_text).value = 'Control'
            table.cell(row=row,column=col_text).style = 'gray_header'

            table.cell(row=row,column=col_l1).value = 'L1'
            table.cell(row=row,column=col_l1).style = 'gray_header'
            table.cell(row=row,column=col_l2).value = 'L2'
            table.cell(row=row,column=col_l2).style = 'gray_header'
            table.cell(row=row,column=col_r).value = 'R'
            table.cell(row=row,column=col_r).style = 'gray_header'

            table.cell(row=row,column=col_link_android).value = 'MSTG Test Coverage'
            table.cell(row=row,column=col_link_android).style = 'gray_header'
            table.merge_cells(start_row=row, end_row=row, start_column=col_link_android, end_column=col_link_ios)

            row = row + 2
        
        # l1 = CHECKMARK if req['L1'] else ""
        # l2 = CHECKMARK if req['L2'] else ""
        # r = CHECKMARK if req['R'] else ""



        # End header

        table.cell(row=row,column=col_id).value = req['id']
        table.cell(row=row,column=col_id).style = 'center'

        table.cell(row=row,column=col_mstg_id).value = mstg_id
        table.cell(row=row,column=col_mstg_id).style = 'center'
        
        table.cell(row=row,column=col_text).value = req['text']
        table.cell(row=row,column=col_text).style = 'text'
        
        if req['L1']:
            # table.cell(row=row,column=col_l1).value = l1
            table.cell(row=row,column=col_l1).style = 'blue'
        if req['L2']:
            # table.cell(row=row,column=col_l2).value = l2
            table.cell(row=row,column=col_l2).style = 'green'
        if req['R']:
            # table.cell(row=row,column=col_r).value = r
            table.cell(row=row,column=col_r).style = 'orange'
        if req.get('links'):
            table.cell(row=row,column=col_link_android).value = get_hyperlink(req['links'][0])
            table.cell(row=row,column=col_link_android).style = 'center'
            if len(req['links']) >= 2:
                table.cell(row=row,column=col_link_ios).value = get_hyperlink(req['links'][1])
                table.cell(row=row,column=col_link_ios).style = 'center'
        else:
            table.cell(row=row,column=col_link_android).value = 'N/A'
            table.cell(row=row,column=col_link_android).style = 'center'
            table.cell(row=row,column=col_link_ios).value = 'N/A'
            table.cell(row=row,column=col_link_ios).style = 'center'
            
        table.row_dimensions[row].height = 55 # points

        row = row+1

    table.sheet_view.showGridLines = False

    # for row in table.iter_rows():
    #     for cell in row:
    #         cell.alignment = align_center

    # for cell in table['D']:
    #     cell.alignment = align_left


    wb.save(filename=output_file)

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Export the MASVS requirements as Excel. Default language is en.')
    parser.add_argument('-m', '--masvs', required=True)
    parser.add_argument('-o', '--outputfile', required=True)

    args = parser.parse_args()

    write_table(args.masvs, args.outputfile)


if __name__ == '__main__':
    main()